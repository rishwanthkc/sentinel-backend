import json
import os
import time
from datetime import datetime, timedelta

import jwt
import requests
from openpyxl import Workbook
from fpdf import FPDF

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BACKEND_DIR = os.path.dirname(BASE_DIR)
INPUT_PATH = os.path.join(BASE_DIR, "input.json")
REPORTS_DIR = os.path.join(BASE_DIR, "reports")
LOGS_DIR = os.path.join(BASE_DIR, "logs")
FINDINGS_DIR = os.path.join(BASE_DIR, "findings")

SAFE_POST_ENDPOINTS = [
    "/auth/register",
    "/auth/login",
    "/emergency/trigger",
    "/emergency/resolve/1",
    "/contacts/add",
    "/reports/submit",
    "/dashboard/resolve/1",
]

ENDPOINTS = [
    {"path": "/", "method": "GET", "role": "Public", "description": "Root health"},
    {"path": "/health", "method": "GET", "role": "Public", "description": "Health status"},
    {"path": "/auth/register", "method": "POST", "role": "Public", "description": "User registration"},
    {"path": "/auth/login", "method": "POST", "role": "Public", "description": "User login"},
    {"path": "/emergency/trigger", "method": "POST", "role": "Public", "description": "Trigger emergency"},
    {"path": "/emergency/history/test@example.com", "method": "GET", "role": "Public", "description": "Emergency history"},
    {"path": "/emergency/resolve/1", "method": "POST", "role": "Public", "description": "Resolve emergency"},
    {"path": "/emergency/active", "method": "GET", "role": "Public", "description": "Active emergencies"},
    {"path": "/contacts/add", "method": "POST", "role": "Public", "description": "Add contact"},
    {"path": "/contacts/test@example.com", "method": "GET", "role": "Public", "description": "Get contacts"},
    {"path": "/reports/submit", "method": "POST", "role": "Public", "description": "Submit report"},
    {"path": "/reports/all", "method": "GET", "role": "Public", "description": "Get all reports"},
    {"path": "/dashboard/stats", "method": "GET", "role": "Public", "description": "Dashboard stats"},
    {"path": "/dashboard/reports", "method": "GET", "role": "Public", "description": "Dashboard reports"},
    {"path": "/dashboard/emergencies", "method": "GET", "role": "Public", "description": "Dashboard emergencies"},
    {"path": "/dashboard/users", "method": "GET", "role": "Public", "description": "Dashboard users"},
    {"path": "/dashboard/analytics", "method": "GET", "role": "Public", "description": "Dashboard analytics"},
    {"path": "/dashboard/hotspots", "method": "GET", "role": "Public", "description": "Dashboard hotspots"},
    {"path": "/dashboard/resolve/1", "method": "POST", "role": "Public", "description": "Resolve emergency from dashboard"},
    {"path": "/safety/risk-points", "method": "GET", "role": "Public", "description": "Risk points"},
    {"path": "/safety/route?origin=0,0&destination=1,1", "method": "GET", "role": "Public", "description": "Safe route"},
]

AUTH_TOKENS = {
    "admin": None,
    "manager": None,
    "user": None,
}

IDOR_PARAMS = ["userId", "accountId", "profileId", "orderId", "transactionId", "reportId"]
INJECTION_PAYLOADS = ["'", "''", "OR 1=1", "' OR '1'='1", '{"$ne":null}', '{"$gt":""}', ";", "&&", "|"]

FINDINGS_JSON = []


def load_input():
    if not os.path.exists(INPUT_PATH):
        raise FileNotFoundError(f"Required input.json not found at {INPUT_PATH}")
    with open(INPUT_PATH, "r", encoding="utf-8") as handle:
        return json.load(handle)


def build_url(base_url, path):
    return base_url.rstrip("/") + path


def record_finding(endpoint, method, role, status, expected, finding, severity, category, note, response_time):
    entry = {
        "endpoint": endpoint,
        "method": method,
        "role": role,
        "status": status,
        "expected_status": expected,
        "finding": finding,
        "severity": severity,
        "response_time_ms": response_time,
        "test_category": category,
        "note": note,
        "timestamp": datetime.utcnow().isoformat() + "Z",
    }
    FINDINGS_JSON.append(entry)
    return entry


def save_json_report():
    path = os.path.join(REPORTS_DIR, "security_report.json")
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(FINDINGS_JSON, handle, indent=2)


def generate_excel_report():
    wb = Workbook()
    ws = wb.active
    ws.title = "Authentication Testing"
    headers = [
        "endpoint",
        "method",
        "role",
        "status",
        "expected_status",
        "finding",
        "severity",
        "test_category",
        "note",
        "response_time_ms",
        "timestamp",
    ]
    ws.append(headers)
    for row in FINDINGS_JSON:
        ws.append([row[h] for h in headers])
    wb.save(os.path.join(REPORTS_DIR, "Backend_Security_Assessment.xlsx"))


def generate_pdf_report():
    pdf = FPDF()
    pdf.set_auto_page_break(True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, "Backend Security Assessment", ln=True, align="C")
    pdf.set_font("Arial", size=12)
    pdf.ln(5)
    pdf.multi_cell(0, 8, "This report contains findings from the generated backend security assessment.")
    pdf.ln(5)
    critical = [f for f in FINDINGS_JSON if f["severity"] == "Critical"]
    high = [f for f in FINDINGS_JSON if f["severity"] == "High"]
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 8, "Critical Findings", ln=True)
    pdf.set_font("Arial", size=12)
    if critical:
        for f in critical[:5]:
            pdf.multi_cell(0, 8, f"{f['endpoint']} [{f['test_category']}] {f['note']}")
            pdf.ln(1)
    else:
        pdf.cell(0, 8, "None detected.", ln=True)
    pdf.ln(5)
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 8, "High Findings", ln=True)
    pdf.set_font("Arial", size=12)
    if high:
        for f in high[:5]:
            pdf.multi_cell(0, 8, f"{f['endpoint']} [{f['test_category']}] {f['note']}")
            pdf.ln(1)
    else:
        pdf.cell(0, 8, "None detected.", ln=True)
    pdf.output(os.path.join(REPORTS_DIR, "Backend_Security_Assessment.pdf"))


def do_rate_limiting_test(base_url):
    burst_count = 30
    url = build_url(base_url, "/health")
    seen_429 = False
    resp = None
    for _ in range(burst_count):
        start = time.time()
        try:
            resp = requests.get(url, timeout=10)
        except Exception as exc:
            record_finding(url, "GET", "N/A", "error", "200", True, "Informational", "Rate Limiting", str(exc), int((time.time() - start) * 1000))
            continue
        seen_429 = seen_429 or resp.status_code == 429
    status_code = resp.status_code if resp is not None else "error"
    if seen_429:
        record_finding(url, "GET", "N/A", status_code, "200", False, "Low", "Rate Limiting", "429 response observed during burst", 0)
    else:
        record_finding(url, "GET", "N/A", status_code, "200", True, "Medium", "Rate Limiting", "No 429 rate limiting detected", 0)


def do_hardcoded_secret_scan():
    paths = [
        os.path.join(os.path.dirname(BASE_DIR), ".env"),
        os.path.join(os.path.dirname(BASE_DIR), ".env.example"),
        os.path.join(os.path.dirname(BASE_DIR), "render.yaml"),
        os.path.join(os.path.dirname(BASE_DIR), "app", "database.py"),
    ]
    patterns = ["password=", "secret=", "apikey=", "token=", "jwt=", "private_key="]
    findings = []
    for path in paths:
        if not os.path.exists(path):
            continue
        with open(path, "r", encoding="utf-8", errors="ignore") as handle:
            for lineno, line in enumerate(handle, start=1):
                if any(p in line.lower() for p in patterns):
                    masked = line.split("=", 1)
                    if len(masked) == 2:
                        findings.append({
                            "file": os.path.relpath(path, BASE_DIR),
                            "line": lineno,
                            "snippet": f"{masked[0]}=********"
                        })
    for item in findings:
        record_finding(item["file"], "N/A", "N/A", "N/A", "N/A", True, "High", "Hardcoded Secret", item["snippet"], 0)


def do_injection_tests(base_url):
    for endpoint in ENDPOINTS:
        if endpoint["method"] != "GET":
            continue
        if endpoint["path"].startswith("/safety/route"):
            record_finding(endpoint["path"], "GET", "N/A", "skipped", "200", False, "Informational", "Injection Detection", "Skipped external Google Directions dependent endpoint.", 0)
            continue
        url = build_url(base_url, endpoint["path"])
        for payload in INJECTION_PAYLOADS:
            start = time.time()
            try:
                resp = requests.get(url, params={"q": payload}, timeout=10)
            except Exception as exc:
                record_finding(url, "GET", "N/A", "error", "200", False, "Informational", "Injection Detection", str(exc), int((time.time() - start) * 1000))
                continue
            note = f"Injection payload {payload} sent. Status {resp.status_code}."
            finding = resp.status_code == 200
            severity = "Low" if finding else "Informational"
            record_finding(url, "GET", "N/A", resp.status_code, "200", finding, severity, "Injection Detection", note, int((time.time() - start) * 1000))


def do_auth_tests(base_url):
    for endpoint in ENDPOINTS:
        url = build_url(base_url, endpoint["path"])
        if endpoint["method"] == "GET":
            try:
                resp = requests.get(url, timeout=10)
            except Exception as exc:
                record_finding(url, "GET", "N/A", "error", "200", False, "Informational", "Authentication Bypass", str(exc), 0)
                continue
            note = "No auth enforcement detected." if resp.status_code == 200 else "Expected auth not present or non-200 response."
            severity = "High" if resp.status_code == 200 else "Low"
            record_finding(url, "GET", "N/A", resp.status_code, "200", resp.status_code == 200, severity, "Authentication Bypass", note, resp.elapsed.microseconds // 1000)
        elif endpoint["method"] == "POST":
            if endpoint["path"] not in SAFE_POST_ENDPOINTS:
                continue
            body = {"test": "value"}
            try:
                resp = requests.post(url, json=body, timeout=10)
            except Exception as exc:
                record_finding(url, "POST", "N/A", "error", "200", False, "Informational", "Authentication Bypass", str(exc), 0)
                continue
            note = "No auth enforcement detected." if resp.status_code == 200 else "Expected auth not present or non-200 response."
            severity = "High" if resp.status_code == 200 else "Low"
            record_finding(url, "POST", "N/A", resp.status_code, "200", resp.status_code == 200, severity, "Authentication Bypass", note, resp.elapsed.microseconds // 1000)


def do_idor_tests(base_url):
    candidate_paths = [
        "/emergency/history/test@example.com",
        "/contacts/test@example.com",
        "/dashboard/resolve/1",
        "/emergency/resolve/1",
    ]
    for path in candidate_paths:
        url = build_url(base_url, path)
        try:
            resp = requests.get(url, timeout=10)
        except Exception as exc:
            record_finding(url, "GET", "N/A", "error", "200", False, "Informational", "IDOR", str(exc), 0)
            continue
        note = f"IDOR check on {url}. Status {resp.status_code}."
        severity = "Critical" if resp.status_code == 200 else "Low"
        record_finding(url, "GET", "N/A", resp.status_code, "200", resp.status_code == 200, severity, "IDOR", note, resp.elapsed.microseconds // 1000)


def run_tests():
    input_data = load_input()
    base_url = input_data["baseUrl"]
    if not base_url.startswith("http"):
        raise ValueError("baseUrl must include http:// or https://")
    if os.path.exists(os.path.join(os.getcwd(), "app", "main.py")):
        os.makedirs(REPORTS_DIR, exist_ok=True)
        do_rate_limiting_test(base_url)
        do_auth_tests(base_url)
        do_idor_tests(base_url)
        do_injection_tests(base_url)
        do_hardcoded_secret_scan()
        save_json_report()
        generate_excel_report()
        generate_pdf_report()
        print("Security assessment complete. Reports generated.")
    else:
        raise RuntimeError("Backend source not found for assessment.")


if __name__ == "__main__":
    run_tests()

import os
import sys
import time
import random
import urllib.request
import urllib.error
import json
from datetime import datetime, timezone
from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Target API Base URL
BASE_URL = "http://127.0.0.1:8000"

# Target endpoints to simulate realistic user behavior
ENDPOINTS = [
    {"name": "Root Check", "path": "/", "method": "GET", "weight": 2},
    {"name": "Health Check", "path": "/health", "method": "GET", "weight": 2},
    {"name": "Active Emergencies", "path": "/emergency/active", "method": "GET", "weight": 3},
    {"name": "Contacts List", "path": "/contacts/user%40sentinel.com", "method": "GET", "weight": 2},
    {"name": "Dashboard Stats", "path": "/dashboard/stats", "method": "GET", "weight": 1}
]

# Flatten endpoints based on weights
ENDPOINT_POOL = []
for ep in ENDPOINTS:
    ENDPOINT_POOL.extend([ep] * ep["weight"])

def send_request(endpoint):
    url = f"{BASE_URL}{endpoint['path']}"
    start_time = time.perf_counter()
    status_code = 0
    error_msg = ""
    
    try:
        req = urllib.request.Request(url, method=endpoint["method"])
        req.add_header("User-Agent", "SentinelLoadTester/1.0")
        
        with urllib.request.urlopen(req, timeout=5) as response:
            status_code = response.getcode()
    except urllib.error.HTTPError as e:
        status_code = e.code
        error_msg = str(e.reason)
    except urllib.error.URLError as e:
        status_code = 503
        error_msg = str(e.reason)
    except Exception as e:
        status_code = 500
        error_msg = str(e)
        
    duration = (time.perf_counter() - start_time) * 1000 # convert to ms
    return {
        "timestamp": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
        "endpoint": endpoint["name"],
        "path": endpoint["path"],
        "status_code": status_code,
        "duration": duration,
        "error": error_msg
    }

def run_load_test(num_users=100, duration_seconds=60):
    print(f"Starting load test with {num_users} virtual users for {duration_seconds} seconds...")
    results = []
    start_time = time.time()
    end_time = start_time + duration_seconds
    
    # Track statistics
    total_requests = 0
    
    def user_loop(user_id):
        nonlocal total_requests
        user_results = []
        while time.time() < end_time:
            # Pick a random endpoint
            ep = random.choice(ENDPOINT_POOL)
            res = send_request(ep)
            res["user_id"] = user_id
            user_results.append(res)
            total_requests += 1
            # Tiny sleep to avoid completely saturating the local thread pool queue scheduler
            time.sleep(0.005)
        return user_results

    with ThreadPoolExecutor(max_workers=num_users) as executor:
        futures = [executor.submit(user_loop, f"User-{i+1}") for i in range(num_users)]
        for fut in as_completed(futures):
            results.extend(fut.result())
            
    print(f"Load test finished. Total requests sent: {len(results)}")
    return results

def save_to_excel(results, num_users, duration_seconds):
    filename = "Sentinel_Load_Test_Report.xlsx"
    wb = openpyxl.Workbook()
    
    # Stylings
    title_font = Font(name="Calibri", size=16, bold=True, color="22D3EE")
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Dark navy slate
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    failed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    # ------------------ SHEET 1: SUMMARY ------------------
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Title
    ws_summary.merge_cells("A1:D1")
    ws_summary["A1"] = "SENTINEL Load Test Execution Summary"
    ws_summary["A1"].font = title_font
    ws_summary["A1"].alignment = center_align
    ws_summary.row_dimensions[1].height = 40
    
    # Basic Config & Overall Metrics
    total_reqs = len(results)
    successful_reqs = len([r for r in results if 200 <= r["status_code"] < 300])
    failed_reqs = total_reqs - successful_reqs
    success_rate = (successful_reqs / total_reqs * 100) if total_reqs > 0 else 0
    
    durations = [r["duration"] for r in results]
    min_time = min(durations) if durations else 0
    max_time = max(durations) if durations else 0
    avg_time = (sum(durations) / len(durations)) if durations else 0
    
    # Calculate 95th Percentile
    sorted_durations = sorted(durations)
    p95_time = sorted_durations[int(len(sorted_durations) * 0.95)] if sorted_durations else 0
    
    rps = total_reqs / duration_seconds
    
    summary_data = [
        [],
        ["Metric", "Value", "Description"],
        ["Test Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Date and time of execution"],
        ["Virtual Users", num_users, "Number of concurrent simulated clients"],
        ["Test Duration", f"{duration_seconds} sec", "Total test run length"],
        ["Total Requests Sent", total_reqs, "Accumulated request volume"],
        ["Successful Requests", successful_reqs, "Requests returning HTTP 2xx status"],
        ["Failed Requests", failed_reqs, "Requests returning errors or timeouts"],
        ["Success Rate", f"{success_rate:.2f}%", "Percentage of successful hits"],
        ["Requests Per Second (RPS)", f"{rps:.2f} req/sec", "Average request throughput"],
        ["Min Response Time", f"{min_time:.1f} ms", "Fastest individual transaction"],
        ["Average Response Time", f"{avg_time:.1f} ms", "Mean transaction speed"],
        ["Max Response Time", f"{max_time:.1f} ms", "Slowest individual transaction"],
        ["95th Percentile Latency", f"{p95_time:.1f} ms", "95% of requests completed under this time"]
    ]
    
    for row_idx, row in enumerate(summary_data, start=2):
        if not row:
            continue
        ws_summary.append(row)
        ws_summary.row_dimensions[row_idx].height = 20
        
        # Style headers
        if row[0] == "Metric":
            for col in range(1, 4):
                cell = ws_summary.cell(row=row_idx, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
        else:
            for col in range(1, 4):
                cell = ws_summary.cell(row=row_idx, column=col)
                cell.border = thin_border
                cell.alignment = center_align if col < 3 else left_align
                if col == 1:
                    cell.font = Font(name="Calibri", size=11, bold=True)
                    
            # Color Success Rate
            if row[0] == "Success Rate":
                cell_val = ws_summary.cell(row=row_idx, column=2)
                cell_val.fill = success_fill if success_rate > 95 else failed_fill
                cell_val.font = Font(name="Calibri", size=11, bold=True)
                
    # ------------------ SHEET 2: ENDPOINT DETAILS ------------------
    ws_endpoints = wb.create_sheet(title="Endpoint Details")
    ws_endpoints.views.sheetView[0].showGridLines = True
    
    ep_headers = ["Endpoint Name", "Path", "Method", "Requests", "RPS", "Min (ms)", "Avg (ms)", "Max (ms)", "Success Rate"]
    ws_endpoints.append(ep_headers)
    ws_endpoints.row_dimensions[1].height = 26
    for col in range(1, len(ep_headers) + 1):
        cell = ws_endpoints.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
    endpoint_names = list(set([r["endpoint"] for r in results]))
    for idx, ep_name in enumerate(endpoint_names, start=2):
        ep_res = [r for r in results if r["endpoint"] == ep_name]
        ep_total = len(ep_res)
        ep_success = len([r for r in ep_res if 200 <= r["status_code"] < 300])
        ep_success_rate = (ep_success / ep_total * 100) if ep_total > 0 else 0
        ep_durations = [r["duration"] for r in ep_res]
        
        ep_min = min(ep_durations) if ep_durations else 0
        ep_max = max(ep_durations) if ep_durations else 0
        ep_avg = (sum(ep_durations) / len(ep_durations)) if ep_durations else 0
        ep_path = ep_res[0]["path"] if ep_res else ""
        ep_method = next((e["method"] for e in ENDPOINTS if e["name"] == ep_name), "GET")
        ep_rps = ep_total / duration_seconds
        
        row_data = [
            ep_name, ep_path, ep_method, ep_total, 
            round(ep_rps, 2), round(ep_min, 1), round(ep_avg, 1), round(ep_max, 1), 
            f"{ep_success_rate:.2f}%"
        ]
        
        ws_endpoints.append(row_data)
        ws_endpoints.row_dimensions[idx].height = 22
        
        for col in range(1, len(row_data) + 1):
            cell = ws_endpoints.cell(row=idx, column=col)
            cell.border = thin_border
            cell.alignment = center_align
            if col == 1:
                cell.font = Font(name="Calibri", size=11, bold=True)
            if col == 9:
                cell.fill = success_fill if ep_success_rate > 95 else failed_fill
                cell.font = Font(name="Calibri", size=11, bold=True)

    # ------------------ SHEET 3: RAW LOGS (SAMPLE) ------------------
    ws_logs = wb.create_sheet(title="Execution Logs")
    ws_logs.views.sheetView[0].showGridLines = True
    
    log_headers = ["No.", "Timestamp (UTC)", "User ID", "Endpoint", "Path", "Status Code", "Duration (ms)", "Status"]
    ws_logs.append(log_headers)
    ws_logs.row_dimensions[1].height = 26
    for col in range(1, len(log_headers) + 1):
        cell = ws_logs.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
    # Sample logs to keep file size reasonable (max 1000 rows)
    sampled_results = results
    if len(results) > 1000:
        sampled_results = random.sample(results, 1000)
        # Sort by timestamp
        sampled_results = sorted(sampled_results, key=lambda x: x["timestamp"])
        
    for idx, r in enumerate(sampled_results, start=2):
        status = "PASSED" if 200 <= r["status_code"] < 300 else f"FAILED ({r['error']})"
        row_data = [
            idx - 1, r["timestamp"], r["user_id"], r["endpoint"], r["path"], r["status_code"], 
            round(r["duration"], 2), status
        ]
        ws_logs.append(row_data)
        ws_logs.row_dimensions[idx].height = 20
        
        for col in range(1, len(row_data) + 1):
            cell = ws_logs.cell(row=idx, column=col)
            cell.border = thin_border
            cell.alignment = center_align
            if col == 8:
                cell.fill = success_fill if "PASSED" in status else failed_fill
                cell.font = Font(name="Calibri", size=11, bold=True, color="006100" if "PASSED" in status else "9C0006")
                
    # Auto-adjust column widths across all sheets
    for ws in [ws_summary, ws_endpoints, ws_logs]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if '\n' in val:
                    lines = val.split('\n')
                    val = max(lines, key=len)
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)
            
    # Save Report
    output_path = os.path.abspath(filename)
    wb.save(output_path)
    print(f"Excel report saved successfully at: {output_path}")
    return output_path

if __name__ == "__main__":
    users = 100
    secs = 60
    
    if len(sys.argv) > 1:
        try:
            users = int(sys.argv[1])
        except ValueError:
            pass
            
    if len(sys.argv) > 2:
        try:
            secs = int(sys.argv[2])
        except ValueError:
            pass
            
    # Run Load Test
    raw_results = run_load_test(num_users=users, duration_seconds=secs)
    
    # Save Results to Excel
    save_to_excel(raw_results, users, secs)
